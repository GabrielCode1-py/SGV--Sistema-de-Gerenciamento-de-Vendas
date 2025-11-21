from flask import Flask, render_template, request, redirect, url_for, flash, session
import os
import pandas as pd
from datetime import datetime, date
import unicodedata
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
app.secret_key = 'chave_secreta_para_flash_e_session_sgv_2025'

# ---------------------------------------------------------------------
# FUNÇÕES AUXILIARES

def normalizar_string(texto):
    """Normaliza strings removendo acentos e convertendo para minúsculas."""
    if not isinstance(texto, str):
        return ""
    return unicodedata.normalize('NFD', texto).encode('ascii', 'ignore').decode('ascii').lower()

def registrar_log(acao):
    """Registra ações no log com timestamp (append)."""
    try:
        with open("vendas_log.txt", "a", encoding="utf-8") as arquivo_log:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            arquivo_log.write(f"[{timestamp}] {acao}\n")
    except Exception as e:
        print(f"Erro ao registrar log: {e}")

def validar_numero_positivo(valor):
    """Valida se um valor numérico é positivo ou zero."""
    try:
        valor_float = float(valor)
        if valor_float < 0:
            return None
        return valor_float
    except (ValueError, TypeError):
        return None

def obter_data_atual():
    """Retorna data atual no formato YYYY-MM-DD."""
    return date.today().isoformat()

def verificar_reset_diario():
    """Verifica se é um novo dia e reseta vendas diárias se necessário."""
    try:
        if os.path.exists("ultima_data.txt"):
            with open("ultima_data.txt", "r") as f:
                ultima_data = f.read().strip()
        else:
            ultima_data = None
        
        data_atual = obter_data_atual()
        
        if ultima_data and ultima_data != data_atual:
            arquivar_vendas_diarias(ultima_data)
            limpar_vendas_diarias()
            registrar_log(f"Reset diário: Vendas de {ultima_data} arquivadas")
        
        with open("ultima_data.txt", "w") as f:
            f.write(data_atual)
    
    except Exception as e:
        registrar_log(f"Erro no reset diário: {str(e)}")

def arquivar_vendas_diarias(data_arquivamento):
    """Arquiva vendas do dia em aba Historico_Vendas."""
    try:
        vendas = carregar_vendas_diarias()
        if not vendas:
            return
        
        df_vendas = pd.DataFrame(vendas)
        
        if os.path.exists("vendas.xlsx"):
            wb = load_workbook("vendas.xlsx")
            if "Historico_Vendas" in wb.sheetnames:
                df_existente = pd.read_excel("vendas.xlsx", sheet_name="Historico_Vendas")
                df_atualizado = pd.concat([df_existente, df_vendas], ignore_index=True)
                del wb["Historico_Vendas"]
            else:
                df_atualizado = df_vendas
            ws = wb.create_sheet("Historico_Vendas")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Historico_Vendas"
            df_atualizado = df_vendas
        
        for r in dataframe_to_rows(df_atualizado, index=False, header=True):
            ws.append(r)
        
        wb.save("vendas.xlsx")
        registrar_log(f"Vendas de {data_arquivamento} arquivadas")
    
    except Exception as e:
        registrar_log(f"Erro ao arquivar vendas: {str(e)}")

def limpar_vendas_diarias():
    """Limpa aba Diario para começar novo dia."""
    try:
        if os.path.exists("vendas.xlsx"):
            wb = load_workbook("vendas.xlsx")
            if "Diario" in wb.sheetnames:
                del wb["Diario"]
            ws = wb.create_sheet("Diario")
            ws.append(["Cliente_ID", "Cliente_Nome", "Produto_Nome", "Tipo_Produto", 
                      "Quantidade_Input", "Quantidade_Total", "Valor_Unitario", 
                      "Valor_Total", "Forma_Pagamento", "Status_Pagamento", "Data"])
            wb.save("vendas.xlsx")
            registrar_log("Vendas diárias resetadas")
    except Exception as e:
        registrar_log(f"Erro ao limpar vendas: {str(e)}")

# ---------------------------------------------------------------------
# MÓDULO: PRODUTOS (estoque.xlsx)

def carregar_produtos():
    """Carrega produtos do arquivo estoque.xlsx."""
    if not os.path.exists("estoque.xlsx"):
        return []
    try:
        df = pd.read_excel("estoque.xlsx")
        if df.empty:
            return []
        produtos = df.to_dict('records')
        return produtos
    except Exception as e:
        registrar_log(f"Erro ao carregar produtos: {str(e)}")
        return []

def salvar_produtos(produtos):
    """Salva produtos no arquivo estoque.xlsx."""
    try:
        if not produtos:
            df = pd.DataFrame(columns=["ID", "Nome", "Tipo", "Valor", "Controlar_Estoque", "Quantidade"])
        else:
            df = pd.DataFrame(produtos)
            df = df[["id", "nome", "tipo", "valor", "controlar_estoque", "quantidade"]].rename(
                columns={"id": "ID", "nome": "Nome", "tipo": "Tipo", "valor": "Valor", 
                        "controlar_estoque": "Controlar_Estoque", "quantidade": "Quantidade"}
            )
        
        if os.path.exists("estoque.xlsx"):
            wb = load_workbook("estoque.xlsx")
            if "Produtos" in wb.sheetnames:
                del wb["Produtos"]
            ws = wb.create_sheet("Produtos")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Produtos"
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        wb.save("estoque.xlsx")
    except Exception as e:
        registrar_log(f"Erro ao salvar produtos: {str(e)}")
        raise

def cadastrar_produto(produtos, nome, tipo, valor, controlar_estoque, quantidade=0):
    """Cadastra novo produto."""
    if not nome or nome.strip() == "":
        return "Erro: Nome vazio!"
    
    tipos_validos = ["unitario", "quilo", "esteira"]
    if tipo not in tipos_validos:
        return "Erro: Tipo inválido!"
    
    valor_validado = validar_numero_positivo(valor)
    if valor_validado is None or valor_validado <= 0:
        return "Erro: Valor inválido!"
    
    if controlar_estoque:
        quantidade_validada = validar_numero_positivo(quantidade)
        if quantidade_validada is None:
            return "Erro: Quantidade inválida!"
    else:
        quantidade_validada = 0
    
    nome_normalizado = normalizar_string(nome)
    for produto in produtos:
        if normalizar_string(produto["nome"]) == nome_normalizado:
            return f"Erro: Produto '{nome}' já existe!"
    
    id = len(produtos) + 1
    produto = {
        "id": id,
        "nome": nome.strip(),
        "tipo": tipo,
        "valor": valor_validado,
        "controlar_estoque": controlar_estoque,
        "quantidade": quantidade_validada
    }
    produtos.append(produto)
    registrar_log(f"Produto cadastrado: {nome}")
    return f"Produto '{nome}' cadastrado!"

def remover_produto(produtos, produto_id):
    """Remove produto."""
    produto = next((p for p in produtos if p["id"] == produto_id), None)
    if not produto:
        return "Erro: Produto não encontrado!"
    produtos.remove(produto)
    salvar_produtos(produtos)
    registrar_log(f"Produto removido: {produto['nome']}")
    return "Produto removido!"

def atualizar_estoque(produtos, produto_id, quantidade_vendida):
    """Atualiza estoque (só se controlar)."""
    produto = next((p for p in produtos if p["id"] == produto_id), None)
    if produto and produto["controlar_estoque"]:
        produto["quantidade"] -= quantidade_vendida
        salvar_produtos(produtos)

# ---------------------------------------------------------------------
# MÓDULO: CLIENTES (vendas.xlsx, aba Clientes) - CORRIGIDO

def carregar_clientes():
    """Carrega clientes do arquivo vendas.xlsx."""
    if not os.path.exists("vendas.xlsx"):
        registrar_log("Arquivo vendas.xlsx não encontrado. Criando novo.")
        return []
    try:
        df = pd.read_excel("vendas.xlsx", sheet_name="Clientes")
        if df.empty:
            return []
        
        colunas_esperadas = ["ID", "Nome", "Telefone", "Observacoes"]
        if not all(col in df.columns for col in colunas_esperadas):
            registrar_log(f"Erro: Colunas inválidas em Clientes. Esperado: {colunas_esperadas}")
            return []
        
        clientes = []
        for _, row in df.iterrows():
            cliente = {
                "id": int(row["ID"]),
                "nome": str(row["Nome"]),
                "telefone": str(row["Telefone"]) if pd.notna(row["Telefone"]) else "",
                "observacoes": str(row["Observacoes"]) if pd.notna(row["Observacoes"]) else ""
            }
            clientes.append(cliente)
        return clientes
    
    except ValueError:
        registrar_log("Aba Clientes não existe. Iniciando com lista vazia.")
        return []
    
    except Exception as e:
        registrar_log(f"Erro ao carregar clientes: {str(e)}")
        return []

def salvar_clientes(clientes):
    """Salva clientes no arquivo vendas.xlsx."""
    try:
        if not clientes:
            df = pd.DataFrame(columns=["ID", "Nome", "Telefone", "Observacoes"])
        else:
            df = pd.DataFrame(clientes)
            for col in ["id", "nome", "telefone", "observacoes"]:
                if col not in df.columns:
                    df[col] = ""
            
            df = df[["id", "nome", "telefone", "observacoes"]].rename(
                columns={"id": "ID", "nome": "Nome", "telefone": "Telefone", "observacoes": "Observacoes"}
            )
        
        if os.path.exists("vendas.xlsx"):
            wb = load_workbook("vendas.xlsx")
            if "Clientes" in wb.sheetnames:
                del wb["Clientes"]
            ws = wb.create_sheet("Clientes")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Clientes"
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        wb.save("vendas.xlsx")
        registrar_log("Clientes salvos")
    
    except Exception as e:
        registrar_log(f"Erro ao salvar clientes: {str(e)}")
        raise Exception(f"Erro ao salvar clientes: {str(e)}")

def cadastrar_cliente(clientes, nome, telefone, observacoes):
    """Cadastra novo cliente."""
    try:
        if not nome or nome.strip() == "":
            return "Erro: Nome vazio!"
        
        nome_normalizado = normalizar_string(nome)
        for cliente in clientes:
            if normalizar_string(cliente.get("nome", "")) == nome_normalizado:
                return f"Erro: Cliente '{nome}' já existe!"
        
        id = len(clientes) + 1
        cliente = {
            "id": id,
            "nome": nome.strip(),
            "telefone": telefone.strip() if telefone else "",
            "observacoes": observacoes.strip() if observacoes else ""
        }
        
        clientes.append(cliente)
        salvar_clientes(clientes)
        registrar_log(f"Cadastrou cliente: {nome}")
        return f"Cliente '{nome}' cadastrado!"
    
    except Exception as e:
        registrar_log(f"Erro ao cadastrar cliente: {str(e)}")
        return f"Erro: {str(e)}"

def remover_cliente(clientes, cliente_id):
    """Remove cliente."""
    try:
        cliente = next((c for c in clientes if c.get("id") == cliente_id), None)
        if not cliente:
            return "Erro: Cliente não encontrado!"
        
        clientes.remove(cliente)
        salvar_clientes(clientes)
        registrar_log(f"Removeu cliente: {cliente.get('nome')}")
        return "Cliente removido!"
    
    except Exception as e:
        registrar_log(f"Erro ao remover cliente: {str(e)}")
        return f"Erro: {str(e)}"

# ---------------------------------------------------------------------
# MÓDULO: VENDAS (vendas.xlsx, aba Diario)

def carregar_vendas_diarias():
    """Carrega vendas diárias (apenas do dia atual)."""
    if not os.path.exists("vendas.xlsx"):
        return []
    try:
        df = pd.read_excel("vendas.xlsx", sheet_name="Diario")
        if df.empty:
            return []
        data_hoje = obter_data_atual()
        vendas = df.to_dict('records')
        vendas_hoje = [v for v in vendas if str(v.get("Data", "")) == data_hoje]
        return vendas_hoje
    except:
        return []

def carregar_historico_vendas():
    """Carrega histórico completo de vendas."""
    if not os.path.exists("vendas.xlsx"):
        return []
    try:
        df = pd.read_excel("vendas.xlsx", sheet_name="Historico_Vendas")
        return df.to_dict('records') if not df.empty else []
    except:
        return []

def salvar_venda_diaria(venda):
    """Salva venda diária."""
    try:
        df_nova = pd.DataFrame([venda])
        
        if os.path.exists("vendas.xlsx"):
            wb = load_workbook("vendas.xlsx")
            if "Diario" in wb.sheetnames:
                df_existente = pd.read_excel("vendas.xlsx", sheet_name="Diario")
                df_atualizado = pd.concat([df_existente, df_nova], ignore_index=True)
                del wb["Diario"]
            else:
                df_atualizado = df_nova
            ws = wb.create_sheet("Diario")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Diario"
            df_atualizado = df_nova
        
        for r in dataframe_to_rows(df_atualizado, index=False, header=True):
            ws.append(r)
        
        wb.save("vendas.xlsx")
    except Exception as e:
        registrar_log(f"Erro ao salvar venda: {str(e)}")
        raise

# ---------------------------------------------------------------------
# MÓDULO: CARRINHO DE VENDAS

def inicializar_carrinho():
    """Inicializa carrinho."""
    if 'carrinho' not in session:
        session['carrinho'] = []
    if 'cliente_id_carrinho' not in session:
        session['cliente_id_carrinho'] = None

def adicionar_item_carrinho(produtos, produto_id, quantidade_input):
    """Adiciona item ao carrinho."""
    produto = next((p for p in produtos if p["id"] == produto_id), None)
    if not produto:
        return None, "Produto não encontrado!"
    
    quantidade_validada = validar_numero_positivo(quantidade_input)
    if not quantidade_validada:
        return None, "Quantidade inválida!"
    
    if produto["tipo"] == "esteira":
        quantidade_total = quantidade_validada * 30
        valor_total = round(quantidade_total * produto["valor"], 2)
        descricao = f"{int(quantidade_validada)} esteiras ({int(quantidade_total)} pães)"
    elif produto["tipo"] == "quilo":
        quantidade_total = quantidade_validada
        valor_total = round(quantidade_validada * produto["valor"], 2)
        descricao = f"{quantidade_validada} kg"
    else:
        quantidade_total = quantidade_validada
        valor_total = round(quantidade_validada * produto["valor"], 2)
        descricao = f"{int(quantidade_validada)} unidades"
    
    if produto["controlar_estoque"] and produto["quantidade"] < quantidade_total:
        return None, f"Estoque insuficiente! Disponível: {produto['quantidade']}"
    
    item = {
        "produto_id": produto_id,
        "produto_nome": produto["nome"],
        "tipo_produto": produto["tipo"],
        "quantidade_input": quantidade_validada,
        "quantidade_total": quantidade_total,
        "valor_unitario": produto["valor"],
        "valor_total": valor_total,
        "descricao": descricao
    }
    return item, None

def finalizar_pedido(clientes, produtos, cliente_id, carrinho, forma_pagamento):
    """Finaliza pedido."""
    cliente = next((c for c in clientes if c["id"] == cliente_id), None)
    if not cliente:
        return "Erro: Cliente não encontrado!"
    
    if not carrinho:
        return "Erro: Carrinho vazio!"
    
    total_pedido = sum(item["valor_total"] for item in carrinho)
    data_venda = obter_data_atual()
    
    for item in carrinho:
        venda = {
            "Cliente_ID": cliente_id,
            "Cliente_Nome": cliente["nome"],
            "Produto_Nome": item["produto_nome"],
            "Tipo_Produto": item["tipo_produto"],
            "Quantidade_Input": item["quantidade_input"],
            "Quantidade_Total": item["quantidade_total"],
            "Valor_Unitario": item["valor_unitario"],
            "Valor_Total": item["valor_total"],
            "Forma_Pagamento": forma_pagamento,
            "Status_Pagamento": "Pago" if forma_pagamento != "pendente" else "Pendente",
            "Data": data_venda
        }
        salvar_venda_diaria(venda)
        atualizar_estoque(produtos, item["produto_id"], item["quantidade_total"])
    
    registrar_log(f"Pedido: {cliente['nome']} - R$ {total_pedido:.2f}")
    return f"Pedido finalizado! Total: R$ {total_pedido:.2f}"

# ---------------------------------------------------------------------
# MÓDULO: FECHAMENTO DE CAIXA

def carregar_fechamentos_caixa():
    """Carrega fechamentos."""
    if not os.path.exists("vendas.xlsx"):
        return []
    try:
        df = pd.read_excel("vendas.xlsx", sheet_name="Fechamento_Caixa")
        return df.to_dict('records') if not df.empty else []
    except:
        return []

def salvar_fechamento_caixa(fechamento):
    """Salva fechamento."""
    try:
        df_novo = pd.DataFrame([fechamento])
        
        if os.path.exists("vendas.xlsx"):
            wb = load_workbook("vendas.xlsx")
            if "Fechamento_Caixa" in wb.sheetnames:
                df_existente = pd.read_excel("vendas.xlsx", sheet_name="Fechamento_Caixa")
                df_atualizado = pd.concat([df_existente, df_novo], ignore_index=True)
                del wb["Fechamento_Caixa"]
            else:
                df_atualizado = df_novo
            ws = wb.create_sheet("Fechamento_Caixa")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Fechamento_Caixa"
            df_atualizado = df_novo
        
        for r in dataframe_to_rows(df_atualizado, index=False, header=True):
            ws.append(r)
        
        wb.save("vendas.xlsx")
        return True
    except Exception as e:
        registrar_log(f"Erro ao salvar fechamento: {str(e)}")
        return False

def processar_fechamento_caixa(data_fechamento, pix, cartao, deposito, dinheiro):
    """Processa fechamento de caixa."""
    try:
        pix = validar_numero_positivo(pix) or 0
        cartao = validar_numero_positivo(cartao) or 0
        deposito = validar_numero_positivo(deposito) or 0
        dinheiro = validar_numero_positivo(dinheiro) or 0
        
        vendas_historico = carregar_historico_vendas()
        vendas_dia = [v for v in vendas_historico if str(v.get("Data", "")) == data_fechamento]
        
        if not vendas_dia:
            vendas_diarias = carregar_vendas_diarias()
            vendas_dia = [v for v in vendas_diarias if str(v.get("Data", "")) == data_fechamento]
        
        if not vendas_dia:
            return None, f"Nenhuma venda em {data_fechamento}!"
        
        total_vendas = sum(float(v.get("Valor_Total", 0)) for v in vendas_dia)
        total_pago = sum(float(v.get("Valor_Total", 0)) for v in vendas_dia 
                        if str(v.get("Status_Pagamento", "")).lower() == "pago")
        total_pendente = total_vendas - total_pago
        
        total_recebido = pix + cartao + deposito + dinheiro
        diferenca = total_recebido - total_pago
        
        fechamento = {
            "Data": data_fechamento,
            "Total_Vendas": round(total_vendas, 2),
            "Total_Pago": round(total_pago, 2),
            "Total_Pendente": round(total_pendente, 2),
            "PIX": round(pix, 2),
            "Cartao": round(cartao, 2),
            "Deposito": round(deposito, 2),
            "Dinheiro": round(dinheiro, 2),
            "Total_Recebido": round(total_recebido, 2),
            "Diferenca": round(diferenca, 2)
        }
        
        if not salvar_fechamento_caixa(fechamento):
            return None, "Erro ao salvar!"
        
        registrar_log(f"Fechamento: {data_fechamento} - R$ {total_recebido:.2f}")
        return fechamento, None
    
    except Exception as e:
        registrar_log(f"Erro no fechamento: {str(e)}")
        return None, f"Erro: {str(e)}"

# ---------------------------------------------------------------------
# MÓDULO: GASTOS

def carregar_gastos():
    """Carrega gastos."""
    if not os.path.exists("gastos.xlsx"):
        return {"fixos": [], "variaveis": []}
    try:
        gastos = {}
        gastos["fixos"] = pd.read_excel("gastos.xlsx", sheet_name="Gastos_Fixos").to_dict('records')
        gastos["variaveis"] = pd.read_excel("gastos.xlsx", sheet_name="Gastos_Variaveis").to_dict('records')
        return gastos
    except:
        return {"fixos": [], "variaveis": []}

def salvar_gastos(gastos):
    """Salva gastos."""
    try:
        wb = Workbook()
        
        ws_fixos = wb.active
        ws_fixos.title = "Gastos_Fixos"
        df_fixos = pd.DataFrame(gastos["fixos"])
        if not df_fixos.empty:
            for r in dataframe_to_rows(df_fixos, index=False, header=True):
                ws_fixos.append(r)
        else:
            ws_fixos.append(["ID", "Descricao", "Valor", "Data_Vencimento"])
        
        ws_variaveis = wb.create_sheet("Gastos_Variaveis")
        df_variaveis = pd.DataFrame(gastos["variaveis"])
        if not df_variaveis.empty:
            for r in dataframe_to_rows(df_variaveis, index=False, header=True):
                ws_variaveis.append(r)
        else:
            ws_variaveis.append(["ID", "Descricao", "Valor", "Quantidade", "Data"])
        
        wb.save("gastos.xlsx")
    except Exception as e:
        registrar_log(f"Erro ao salvar gastos: {str(e)}")
        raise

def cadastrar_gasto(gastos, tipo_gasto, descricao, valor, data_vencimento=None, quantidade=None):
    """Cadastra gasto."""
    if not descricao or descricao.strip() == "":
        return "Erro: Descrição vazia!"
    
    valor_validado = validar_numero_positivo(valor)
    if valor_validado is None or valor_validado <= 0:
        return "Erro: Valor inválido!"
    
    if tipo_gasto == "fixo":
        lista = gastos["fixos"]
        id = len(lista) + 1
        gasto = {
            "ID": id,
            "Descricao": descricao.strip(),
            "Valor": valor_validado,
            "Data_Vencimento": data_vencimento or ""
        }
    else:
        lista = gastos["variaveis"]
        id = len(lista) + 1
        quantidade_validada = validar_numero_positivo(quantidade) if quantidade else 1
        gasto = {
            "ID": id,
            "Descricao": descricao.strip(),
            "Valor": valor_validado,
            "Quantidade": quantidade_validada or 1,
            "Data": obter_data_atual()
        }
    
    lista.append(gasto)
    salvar_gastos(gastos)
    registrar_log(f"Gasto cadastrado: {descricao}")
    return "Gasto cadastrado!"

def remover_gasto(gastos, tipo_gasto, gasto_id):
    """Remove gasto."""
    lista = gastos["fixos"] if tipo_gasto == "fixo" else gastos["variaveis"]
    gasto = next((g for g in lista if g["ID"] == gasto_id), None)
    if not gasto:
        return "Erro: Gasto não encontrado!"
    lista.remove(gasto)
    salvar_gastos(gastos)
    registrar_log(f"Gasto removido: {gasto['Descricao']}")
    return "Gasto removido!"

# ---------------------------------------------------------------------
# MÓDULO: FECHAMENTO MENSAL

def fechamento_mensal():
    """Cria fechamento mensal."""
    vendas = carregar_historico_vendas()
    gastos_data = carregar_gastos()
    
    if not vendas:
        return "Nenhuma venda para fechar!"
    
    mes_atual = date.today().strftime("%Y_%m")
    nome_aba = f"Mes_{mes_atual}"
    
    try:
        if os.path.exists("vendas.xlsx"):
            wb = load_workbook("vendas.xlsx")
            if nome_aba in wb.sheetnames:
                return f"Aba '{nome_aba}' já existe!"
            wb.close()
    except:
        pass
    
    vendas_mes = [v for v in vendas if str(v.get("Data", "")).startswith(mes_atual.replace("_", "-"))]
    
    if not vendas_mes:
        return "Nenhuma venda neste mês!"
    
    df_vendas = pd.DataFrame(vendas_mes)
    total_vendas = df_vendas["Valor_Total"].sum()
    
    total_gastos_fixos = sum(g["Valor"] for g in gastos_data["fixos"])
    total_gastos_variaveis = sum(g["Valor"] * g.get("Quantidade", 1) for g in gastos_data["variaveis"])
    total_gastos = total_gastos_fixos + total_gastos_variaveis
    
    lucro = total_vendas - total_gastos
    
    dados_resumo = [
        {"Descricao": "Total Vendas", "Valor": round(total_vendas, 2)},
        {"Descricao": "Gastos Fixos", "Valor": round(total_gastos_fixos, 2)},
        {"Descricao": "Gastos Variáveis", "Valor": round(total_gastos_variaveis, 2)},
        {"Descricao": "Total Gastos", "Valor": round(total_gastos, 2)},
        {"Descricao": "Lucro", "Valor": round(lucro, 2)}
    ]
    
    df_resumo = pd.DataFrame(dados_resumo)
    
    try:
        wb = load_workbook("vendas.xlsx")
        ws = wb.create_sheet(nome_aba)
        for r in dataframe_to_rows(df_resumo, index=False, header=True):
            ws.append(r)
        wb.save("vendas.xlsx")
        registrar_log(f"Fechamento mensal: {nome_aba}")
        return f"Fechamento criado: {nome_aba} - Lucro: R$ {lucro:.2f}"
    except Exception as e:
        return f"Erro: {str(e)}"

# ---------------------------------------------------------------------
# ROTAS DO FLASK

verificar_reset_diario()

produtos = carregar_produtos()
clientes = carregar_clientes()
gastos = carregar_gastos()
registrar_log("Sistema iniciado")

@app.before_request
def before_request():
    """Verifica reset diário antes de cada requisição."""
    verificar_reset_diario()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/cadastrar_produto', methods=['GET', 'POST'])
def cadastrar_produto_route():
    try:
        if request.method == 'POST':
            nome = request.form.get('nome')
            tipo = request.form.get('tipo')
            valor = request.form.get('valor')
            controlar_estoque = request.form.get('controlar_estoque') == 'sim'
            quantidade = request.form.get('quantidade', 0) if controlar_estoque else 0
            mensagem = cadastrar_produto(produtos, nome, tipo, valor, controlar_estoque, quantidade)
            flash(mensagem)
            return redirect(url_for('index'))
        return render_template('cadastrar_produto.html')
    except Exception as e:
        registrar_log(f"Erro na rota cadastrar_produto: {str(e)}")
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/remover_produto/<int:produto_id>', methods=['POST'])
def remover_produto_route(produto_id):
    try:
        mensagem = remover_produto(produtos, produto_id)
        flash(mensagem)
        return redirect(url_for('listar_produtos_route'))
    except Exception as e:
        flash(f"Erro: {str(e)}")
        return redirect(url_for('listar_produtos_route'))

@app.route('/produtos')
def listar_produtos_route():
    try:
        return render_template('produtos.html', produtos=produtos)
    except Exception as e:
        registrar_log(f"Erro na rota produtos: {str(e)}")
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/cadastrar_cliente', methods=['GET', 'POST'])
def cadastrar_cliente_route():
    try:
        if request.method == 'POST':
            nome = request.form.get('nome', '')
            telefone = request.form.get('telefone', '')
            observacoes = request.form.get('observacoes', '')
            
            mensagem = cadastrar_cliente(clientes, nome, telefone, observacoes)
            flash(mensagem)
            return redirect(url_for('index'))
        
        return render_template('cadastrar_cliente.html')
    
    except Exception as e:
        registrar_log(f"Erro na rota cadastrar_cliente: {str(e)}")
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/remover_cliente/<int:cliente_id>', methods=['POST'])
def remover_cliente_route(cliente_id):
    try:
        mensagem = remover_cliente(clientes, cliente_id)
        flash(mensagem)
        return redirect(url_for('listar_clientes_route'))
    except Exception as e:
        registrar_log(f"Erro ao remover cliente: {str(e)}")
        flash(f"Erro: {str(e)}")
        return redirect(url_for('listar_clientes_route'))

@app.route('/clientes')
def listar_clientes_route():
    try:
        clientes_atualizados = carregar_clientes()
        return render_template('clientes.html', clientes=clientes_atualizados)
    except Exception as e:
        registrar_log(f"Erro na rota listar_clientes: {str(e)}")
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/cliente/<int:cliente_id>')
def cliente_detalhes(cliente_id):
    try:
        cliente = next((c for c in clientes if c["id"] == cliente_id), None)
        if not cliente:
            flash("Cliente não encontrado!")
            return redirect(url_for('listar_clientes_route'))
        
        inicializar_carrinho()
        session['cliente_id_carrinho'] = cliente_id
        
        vendas_cliente = [v for v in carregar_vendas_diarias() if v.get("Cliente_ID") == cliente_id]
        total_carrinho = sum(item["valor_total"] for item in session.get('carrinho', []))
        
        return render_template('cliente_detalhes.html', 
                             cliente=cliente, 
                             produtos=produtos, 
                             carrinho=session.get('carrinho', []),
                             total_carrinho=total_carrinho,
                             vendas_cliente=vendas_cliente)
    except Exception as e:
        registrar_log(f"Erro na rota cliente_detalhes: {str(e)}")
        flash(f"Erro: {str(e)}")
        return redirect(url_for('listar_clientes_route'))

@app.route('/adicionar_carrinho', methods=['POST'])
def adicionar_carrinho():
    try:
        inicializar_carrinho()
        cliente_id = session.get('cliente_id_carrinho')
        if not cliente_id:
            flash("Selecione um cliente!")
            return redirect(url_for('listar_clientes_route'))
        
        produto_id = int(request.form.get('produto_id'))
        quantidade = request.form.get('quantidade')
        
        item, erro = adicionar_item_carrinho(produtos, produto_id, quantidade)
        if erro:
            flash(erro)
        else:
            session['carrinho'].append(item)
            session.modified = True
            flash(f"Adicionado: {item['descricao']} - R$ {item['valor_total']:.2f}")
        
        return redirect(url_for('cliente_detalhes', cliente_id=cliente_id))
    except Exception as e:
        registrar_log(f"Erro ao adicionar carrinho: {str(e)}")
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/finalizar_pedido', methods=['POST'])
def finalizar_pedido_route():
    try:
        cliente_id = session.get('cliente_id_carrinho')
        carrinho = session.get('carrinho', [])
        forma_pagamento = request.form.get('forma_pagamento', 'pendente')
        
        if not cliente_id:
            flash("Nenhum cliente selecionado!")
            return redirect(url_for('listar_clientes_route'))
        
        mensagem = finalizar_pedido(clientes, produtos, cliente_id, carrinho, forma_pagamento)
        flash(mensagem)
        
        session['carrinho'] = []
        session['cliente_id_carrinho'] = None
        session.modified = True
        
        return redirect(url_for('index'))
    except Exception as e:
        registrar_log(f"Erro ao finalizar pedido: {str(e)}")
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/limpar_carrinho', methods=['POST'])
def limpar_carrinho():
    try:
        cliente_id = session.get('cliente_id_carrinho')
        session['carrinho'] = []
        session.modified = True
        flash("Carrinho limpo!")
        if cliente_id:
            return redirect(url_for('cliente_detalhes', cliente_id=cliente_id))
        return redirect(url_for('index'))
    except Exception as e:
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/fechamento_caixa', methods=['GET', 'POST'])
def fechamento_caixa_route():
    try:
        if request.method == 'POST':
            data_fechamento = request.form.get('data_fechamento')
            pix = request.form.get('pix', 0)
            cartao = request.form.get('cartao', 0)
            deposito = request.form.get('deposito', 0)
            dinheiro = request.form.get('dinheiro', 0)
            
            fechamento, erro = processar_fechamento_caixa(data_fechamento, pix, cartao, deposito, dinheiro)
            if erro:
                flash(erro)
            else:
                diferenca = fechamento['Diferenca']
                if diferenca < 0:
                    flash(f"Fechamento realizado! FALTA: R$ {abs(diferenca):.2f}")
                elif diferenca > 0:
                    flash(f"Fechamento realizado! SOBRA: R$ {diferenca:.2f}")
                else:
                    flash("Fechamento realizado! Valores conferem!")
            
            return redirect(url_for('fechamento_caixa_route'))
        
        vendas = carregar_vendas_diarias()
        fechamentos = carregar_fechamentos_caixa()
        data_hoje = obter_data_atual()
        
        total_dia = sum(v.get("Valor_Total", 0) for v in vendas)
        total_pago_dia = sum(v.get("Valor_Total", 0) for v in vendas if v.get("Status_Pagamento") == "Pago")
        total_pendente_dia = total_dia - total_pago_dia
        
        return render_template('fechamento_caixa.html', 
                             vendas=vendas, 
                             fechamentos=fechamentos,
                             data_hoje=data_hoje,
                             total_dia=total_dia,
                             total_pago_dia=total_pago_dia,
                             total_pendente_dia=total_pendente_dia)
    
    except Exception as e:
        registrar_log(f"Erro na rota fechamento_caixa: {str(e)}")
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/gastos', methods=['GET', 'POST'])
def gastos_route():
    try:
        if request.method == 'POST':
            tipo_gasto = request.form.get('tipo_gasto')
            descricao = request.form.get('descricao')
            valor = request.form.get('valor')
            data_vencimento = request.form.get('data_vencimento')
            quantidade = request.form.get('quantidade')
            mensagem = cadastrar_gasto(gastos, tipo_gasto, descricao, valor, data_vencimento, quantidade)
            flash(mensagem)
            return redirect(url_for('gastos_route'))
        
        total_fixos = sum(g["Valor"] for g in gastos["fixos"])
        total_variaveis = sum(g["Valor"] * g.get("Quantidade", 1) for g in gastos["variaveis"])
        
        return render_template('gastos.html', 
                             gastos=gastos, 
                             total_fixos=total_fixos,
                             total_variaveis=total_variaveis)
    except Exception as e:
        registrar_log(f"Erro na rota gastos: {str(e)}")
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/remover_gasto/<tipo_gasto>/<int:gasto_id>', methods=['POST'])
def remover_gasto_route(tipo_gasto, gasto_id):
    try:
        mensagem = remover_gasto(gastos, tipo_gasto, gasto_id)
        flash(mensagem)
        return redirect(url_for('gastos_route'))
    except Exception as e:
        flash(f"Erro: {str(e)}")
        return redirect(url_for('gastos_route'))

@app.route('/relatorios')
def relatorios():
    try:
        vendas_diarias = carregar_vendas_diarias()
        total_geral = sum(v.get("Valor_Total", 0) for v in vendas_diarias)
        
        vendas_por_cliente = {}
        for venda in vendas_diarias:
            cliente = venda.get("Cliente_Nome", "Desconhecido")
            if cliente not in vendas_por_cliente:
                vendas_por_cliente[cliente] = {"vendas": [], "total": 0}
            vendas_por_cliente[cliente]["vendas"].append(venda)
            vendas_por_cliente[cliente]["total"] += venda.get("Valor_Total", 0)
        
        return render_template('relatorios.html', 
                             vendas_diarias=vendas_diarias, 
                             total_geral=total_geral,
                             vendas_por_cliente=vendas_por_cliente)
    except Exception as e:
        registrar_log(f"Erro na rota relatorios: {str(e)}")
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/fechamento_mensal')
def fechamento_mensal_route():
    try:
        mensagem = fechamento_mensal()
        flash(mensagem)
        return redirect(url_for('index'))
    except Exception as e:
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/listar')
def listar():
    try:
        return render_template('listar.html', produtos=produtos)
    except Exception as e:
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

@app.route('/salvar')
def salvar():
    try:
        salvar_produtos(produtos)
        salvar_clientes(clientes)
        salvar_gastos(gastos)
        return render_template('salvar.html')
    except Exception as e:
        flash(f"Erro: {str(e)}")
        return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=False, host='127.0.0.1', port=5000)